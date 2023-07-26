# -*- coding: utf-8 -*-
from itertools import cycle
from functools import partial
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from operator import attrgetter
from Products.Five.browser import BrowserView
from io import BytesIO
from zope.component import getUtility
from zope.schema.interfaces import IVocabularyFactory
from emrt.necd.content.vocabularies.vocabularies import get_registry_interface_field_data
from emrt.necd.content.vocabularies.vocabularies import INECDVocabularies


XLS_SAMPLE_HEADER_INVENTORY = (
    'Observation description', 'Country', 'NFR code',
    'Inventory year', 'Pollutants', 'Review year', 'Fuel', 'MS key category',
    'Parameter', 'Description flags', 'Initial question text',
)

XLS_SAMPLE_HEADER_PROJECTION = (
    'Observation description', 'Country', 'NFR code projection',
    # 'NFR inventories category code',
    'Projection year', 'Reference year',
    'Pollutants', 'Scenario type', 'Review year', 'Activity data type',
    'Activity data', 'MS Key Category', 'Parameter', 'Description Flags',
    'Initial question text',
)

DESC = 'Description of the observation'
NFR_CODE = '1A1'
INVENTORY_YEAR = '2018'
REVIEW_YEAR = '2018'
REFERENCE_YEAR = '2018'
QUESTION_TEXT = (
    'The text of an initial Q&A question. '
    'Leave empty if you do not wish to add an initial question.'
)


# UnicodeEncodeError
def decode(s):
    return s.decode('UTF-8', 'replace') if isinstance(s, str) else s


def _get_vocabulary(context, name):
    factory = getUtility(IVocabularyFactory, name=name)
    return factory(context)


def adjust_column_widths(worksheet):
    """Resize columns to match header text width."""
    widths: T.Dict[str, int] = defaultdict(int)

    for row in worksheet:
        for col, cell in enumerate(row, 1):
            if type(cell).__name__ == "MergedCell":
                continue  # ignore
            letter = get_column_letter(col)
            value = str(cell.value)
            width = len(value)
            if width > widths[letter]:
                widths[letter] = width

    for letter, width in widths.items():
        worksheet.column_dimensions[letter].width = width


def adjust_row_heights(worksheet):
    for idx, row in enumerate(worksheet, 1):
        for col, cell in enumerate(row, 1):
            if type(cell).__name__ == "MergedCell":
                continue  # ignore
            letter = get_column_letter(col)
            # calculate height based on text length compared to columnwidth
            width = int(worksheet.column_dimensions[letter].width)
            text_length = len(str(cell.value)) if cell.value is not None else 0
            lines_estimate = (round(text_length / width) + 1) + str(
                cell.value
            ).count("\n")
            calculated_height = (
                round(lines_estimate * cell.font.size) if lines_estimate else 0
            )
            # update height if needed (default is 12.75pt)
            current_height = worksheet.row_dimensions[idx].height or 20
            if current_height < calculated_height:
                worksheet.row_dimensions[idx].height = calculated_height


class GetSampleXLS(BrowserView):

    def populate_cells(self, sheet):
        is_projection = self.context.type == 'projection'
        get_vocabulary = partial(_get_vocabulary, self.context)
        get_title = attrgetter('title')

        if is_projection:
            header = XLS_SAMPLE_HEADER_PROJECTION
            proj_years = ['2025', '2030', '2040', '2050']
            scenario_voc = get_vocabulary('emrt.necd.content.scenario_type')
            act_type_v = get_vocabulary('emrt.necd.content.activity_data_type')

            p_year = '\n'.join(proj_years)
            # nfr_inventories = cycle([NFR_CODE, None])
            act_type = cycle(list(map(attrgetter('value'), act_type_v)) + [None])
            activity_data = get_registry_interface_field_data(
                INECDVocabularies, 'activity_data')
            scenario = cycle(['\n'.join(map(get_title, scenario_voc)), None])
        else:
            header = XLS_SAMPLE_HEADER_INVENTORY
            fuel_voc = get_vocabulary('emrt.necd.content.fuel')
            # not a mandatory field, value can be none
            fuels = cycle(list(map(get_title, fuel_voc)) + [None])

        country_voc = get_vocabulary('emrt.necd.content.eea_member_states')
        pollutants_voc = get_vocabulary('emrt.necd.content.pollutants')
        parameter_voc = get_vocabulary('emrt.necd.content.parameter')
        description_flags_voc = get_vocabulary('emrt.necd.content.highlight')

        countries = list(map(get_title, country_voc))

        ms_key_categ = cycle(['True', None])
        pollutants = '\n'.join(map(get_title, pollutants_voc))
        parameter = '\n'.join(map(get_title, parameter_voc))
        description_flags = cycle(
            ['\n'.join(map(get_title, description_flags_voc)), None]
        )

        sheet.append(header)
        for idx, country in enumerate(countries):
            # get a value based on the country index position
            ms_key_cat = next(ms_key_categ)
            desc_fl = next(description_flags)

            if is_projection:
                # nfr_i = next(nfr_inventories)
                activity_type = next(act_type)
                activity = (
                    '\n'.join(activity_data[activity_type]).encode('utf-8')
                    if activity_type else None
                )
                scenario_type = next(scenario)
                row = [DESC, country, NFR_CODE,
                       # nfr_i,
                       p_year, REFERENCE_YEAR,
                       pollutants, scenario_type, REVIEW_YEAR, activity_type,
                       decode(activity), ms_key_cat, parameter, desc_fl,
                       QUESTION_TEXT]
            else:
                fuel = next(fuels)
                row = [DESC, country, NFR_CODE, INVENTORY_YEAR, pollutants,
                       REVIEW_YEAR, fuel, ms_key_cat, parameter, desc_fl,
                       QUESTION_TEXT]

            sheet.append(row)

    def __call__(self):
        wb = Workbook()
        sheet = wb.create_sheet('Observation', 0)

        self.populate_cells(sheet)

        adjust_column_widths(sheet)
        adjust_row_heights(sheet)

        xls = BytesIO()

        wb.save(xls)

        xls.seek(0)
        filename = 'observation_import_sample.xlsx'
        self.request.response.setHeader(
            'Content-type', 'application/vnd.ms-excel; charset=utf-8'
        )
        self.request.response.setHeader(
            'Content-Disposition', 'attachment; filename={0}'.format(filename)
        )
        return xls.read()
