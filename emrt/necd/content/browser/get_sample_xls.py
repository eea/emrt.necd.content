from functools import partial
from openpyxl import Workbook
from openpyxl.styles import Alignment
from operator import attrgetter
from Products.Five.browser import BrowserView
from StringIO import StringIO
from zope.component import getUtility
from zope.schema.interfaces import IVocabularyFactory


XLS_SAMPLE_HEADER = [
    'Observation description', 'Country', 'NFR Code',
    'Inventory Year', 'Pollutants', 'Review Year', 'Fuel', 'MS Key Category',
    'Parameter'
]

DESC = 'Description of the observation'
NFR_CODE = '1A1'
INVENTORY_YEAR = '2017'
REVIEW_YEAR = '2017'


def _get_vocabulary(context, name):
    factory = getUtility(IVocabularyFactory, name=name)
    return factory(context)

def get_fuel_val(fuels, country_index):
    """Generate all possible values for fuels"""
    fuels_len = len(fuels)

    if country_index < fuels_len:
        return fuels[country_index]
    return fuels[country_index%fuels_len]


class GetSampleXLS(BrowserView):

    def populate_cells(self, sheet):

        get_vocabulary = partial(_get_vocabulary, self.context)
        get_title = attrgetter('title')

        country_voc = get_vocabulary('emrt.necd.content.eea_member_states')
        pollutants_voc = get_vocabulary('emrt.necd.content.pollutants')
        parameter_voc = get_vocabulary('emrt.necd.content.parameter')
        fuel_voc = get_vocabulary('emrt.necd.content.fuel')

        countries = map(get_title, country_voc)
        fuels = map(get_title, fuel_voc)
        #not a mandatory field, value can be none
        fuels.append(None)
        ms_key_categ = ['True', None]
        pollutants = '\n'.join(map(get_title, pollutants_voc))
        parameter = '\n'.join(map(get_title, parameter_voc))

        sheet.append(XLS_SAMPLE_HEADER)
        for idx, country in enumerate(countries):
            fuel = get_fuel_val(fuels, idx)

            #get a value based on the country index position
            ms_key_cat = ms_key_categ[idx%2]
            row = [DESC, country, NFR_CODE, INVENTORY_YEAR, pollutants,
                   REVIEW_YEAR, fuel, ms_key_cat, parameter]
            sheet.append(row)

    def __call__(self):
        wb = Workbook()
        sheet = wb.create_sheet('Observation', 0)

        self.populate_cells(sheet)

        # wrap text for multi line cells
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

        # set cell max width
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column].width = length

        xls = StringIO()

        wb.save(xls)

        xls.seek(0)
        filename = 'observation_import_sample.xlsx'
        self.request.response.setHeader(
            'Content-type', 'application/vnd.ms-excel; charset=utf-8'
        )
        self.request.response.setHeader(
            'Content-Disposition', 'attachment; filename={0}'.format(filename)
        )
        return xls.read()
